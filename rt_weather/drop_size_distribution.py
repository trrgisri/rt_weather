#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
@file drop_size_distribution.py
@brief Raindrop size distribution acquired by a disdrometer

@author Yasushi SUMI <y.sumi@aist.go.jp>

Copyright (C) 2023 AIST
Released under the MIT license
https://opensource.org/licenses/mit-license.php
"""
# %%
import os
from openpyxl import load_workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.label import DataLabel, DataLabelList
from openpyxl.utils.units import pixels_to_EMU
import numpy as np
import pandas as pd
import pandas.tseries.offsets as offsets
import matplotlib.pyplot as plt
import seaborn as sns
import tempfile

# %%
class DropSizeDistribution(object):
    ##
    # @brief Constructor
    #
    # @param dict_psvd             Dictionary with datetime as Key and PSVD matrix
    #                                as Value in a Pandas DataFrame with particle
    #                                diameters as Index and falling velocities as Column.
    # @param sampling_area         Sampling area of the disdrometer [mm^2]
    # @param sampling_interval     Sampling interval of the disdrometer [sec]
    # @param diameter_binwidth     List of the raindrop diameter binwidths of the PSVD matrix
    # @param rate_rainfall         Rate of rainfall [mm/h];
    #                                if None, unknown or varied
    # @param start_stable_timestr  Start time of stable rainfall in "HH:MM:SS";
    #                                if None, same as start_timestr
    # @param end_stable_timestr    End time of stable rainfall in "HH:MM:SS";
    #                                if None, same as end_timestr
    # @param start_offset_min      Offset of the start time in minutes
    # @param end_offset_min        Offset of the end time in minutes,
    #                                usually a negative value.
    # @param raindrop_diameters    List of classes of raindrop diameters
    # @param raindrop_velocities   List of classes of raindrop velocities
    """ Example of the datetime and the PSVD matrix
Timestamp('2022-12-09 11:41:00'),
Speed     20.80  17.60  15.20  13.60  12.00  10.40  8.80   7.60   6.80   6.00   5.20   4.40   3.80   ...  1.50   1.30   1.10   0.95   0.85   0.75   0.65   0.55   0.45   0.35   0.25   0.15   0.05 
Diameter                                                                                             ...                                                                                           
0.062         0      0      0      0      0      0      0      0      0      0      0      0      0  ...      0      0      0      0      0      0      0      0      0      0      0      0      0
0.187         0      0      0      0      0      0      0      0      0      0      0      0      0  ...      0      0      0      0      0      0      0      0      0      0      0      0      0
0.312         0      0      0      0      0      0      0      0      0      0      0      0      0  ...      1      0      1      0      1      0      0      0      0      0      0      0      0
0.437         0      0      0      0      0      0      0      0      0      0      0      0      0  ...      2      2      1      0      0      1      1      0      0      0      0      0      0
0.562         0      0      0      0      0      0      0      0      0      0      0      0      0  ...      3      0      0      0      0      0      0      0      0      0      0      0      0
0.687         0      0      0      0      0      0      0      0      0      0      0      0     24  ...      4      0      0      0      0      0      0      0      0      0      0      0      0
0.812         0      0      0      0      0      0      0      0      0      0      0      6      8  ...      5      1      1      0      1      1      0      0      0      0      0      0      0
0.937         0      0      0      0      0      0      0      0      0      0      0      0     14  ...      5      1      0      0      0      0      0      0      0      0      0      0      0
1.062         0      0      0      0      0      0      0      0      0      0      0      2     17  ...      1      0      0      0      0      0      0      0      0      0      0      0      0
1.187         0      0      0      0      0      0      0      0      0      0      0      2     34  ...      0      0      0      0      0      0      0      0      0      0      0      0      0
1.375         0      0      0      0      0      0      0      0      0      0      0     15     19  ...      0      0      0      0      0      0      0      0      0      0      0      0      0
1.625         0      0      0      0      0      0      0      0      0      0      0     16      9  ...      0      0      0      0      0      0      0      0      0      0      0      0      0
1.875         0      0      0      0      0      0      0      0      0      0      6      5      2  ...      0      0      0      0      0      0      0      0      0      0      0      0      0
2.125         0      0      0      0      0      0      0      0      0      0      1      1      0  ...      0      0      0      0      0      0      0      0      0      0      0      0      0
2.375         0      0      0      0      0      0      0      0      1      1      1      0      0  ...      0      0      0      0      0      0      0      0      0      0      0      0      0
2.750         0      0      0      0      0      0      0      0      2      0      0      0      0  ...      0      0      0      0      0      0      0      0      0      0      0      0      0
3.250         0      0      0      0      0      0      0      1      0      0      0      0      0  ...      0      0      0      0      0      0      0      0      0      0      0      0      0
3.750         0      0      0      0      0      0      0      0      0      0      0      0      0  ...      0      0      0      0      0      0      0      0      0      0      0      0      0
4.250         0      0      0      0      0      0      0      0      0      0      0      0      0  ...      0      0      0      0      0      0      0      0      0      0      0      0      0
4.750         0      0      0      0      0      0      0      0      0      0      0      0      0  ...      0      0      0      0      0      0      0      0      0      0      0      0      0
5.500         0      0      0      0      0      0      0      0      0      0      0      0      0  ...      0      0      0      0      0      0      0      0      0      0      0      0      0
6.500         0      0      0      0      0      0      0      0      0      0      0      0      0  ...      0      0      0      0      0      0      0      0      0      0      0      0      0
7.500         0      0      0      0      0      0      0      0      0      0      0      0      0  ...      0      0      0      0      0      0      0      0      0      0      0      0      0
8.500         0      0      0      0      0      0      0      0      0      0      0      0      0  ...      0      0      0      0      0      0      0      0      0      0      0      0      0
9.500         0      0      0      0      0      0      0      0      0      0      0      0      0  ...      0      0      0      0      0      0      0      0      0      0      0      0      0
11.000        0      0      0      0      0      0      0      0      0      0      0      0      0  ...      0      0      0      0      0      0      0      0      0      0      0      0      0
13.000        0      0      0      0      0      0      0      0      0      0      0      0      0  ...      0      0      0      0      0      0      0      0      0      0      0      0      0
15.000        0      0      0      0      0      0      0      0      0      0      0      0      0  ...      0      0      0      0      0      0      0      0      0      0      0      0      0
17.000        0      0      0      0      0      0      0      0      0      0      0      0      0  ...      0      0      0      0      0      0      0      0      0      0      0      0      0
19.000        0      0      0      0      0      0      0      0      0      0      0      0      0  ...      0      0      0      0      0      0      0      0      0      0      0      0      0
21.500        0      0      0      0      0      0      0      0      0      0      0      0      0  ...      0      0      0      0      0      0      0      0      0      0      0      0      0
24.500        0      0      0      0      0      0      0      0      0      0      0      0      0  ...      0      0      0      0      0      0      0      0      0      0      0      0      0
    """
    def __init__(self, dict_psvd,
                        sampling_area,
                        sampling_interval,
                        diameter_binwidth,
                        rate_rainfall=None,
                        start_stable_timestr=None,
                        end_stable_timestr=None,
                        start_offset_min=0,
                        end_offset_min=0,
                        dict_precip=None,
                        show_heatmap_p=False,
                        #min_observable_diameter=0.0,
                        ):
        if rate_rainfall == None and dict_precip == None:
            raise ValueError(f'>>> rate_rainfall and dict_precip are both None.')

        self.raindrop_diameters, self.raindrop_velocities = self._extract_dims_velocs(dict_psvd)
        print(self.raindrop_diameters)
        #self.raindrop_diameter_differences = self.calc_diameter_differences(self.raindrop_diameters)
        if len(self.raindrop_diameters) != len(diameter_binwidth):
            raise ValueError(f'>>> Bad length of raindrop_diameter_binwidth: {len(diameter_binwidth)} for {len(self.raindrop_diameters)} expected.')

        self.raindrop_diameter_binwidth = diameter_binwidth
        '''
        if min_observable_diameter > 0.0:
            if min_observable_diameter < self.raindrop_diameter_differences[0]:
                self.raindrop_diameter_differences -= min_diameter
            else:
                raise ValueError('>>> Minimum observable diameter out of range')
        '''

        self.sampling_area = sampling_area
        self.sampling_interval = sampling_interval

        # list of DataFrame for PSVD (Particle Size-Velocity Distribution)
        self.dict_psvd_all = dict_psvd

        # Set stable rainfall times
        self._set_stable_times(start_stable_timestr,
                                end_stable_timestr,
                                start_offset_min,
                                end_offset_min)

        # stable timeseries data
        # _dft_stable = self.df_timeseries_all.T[self.start_stable_time : self.end_stable_time]
        # self.df_timeseries_stable = _dft_stable.T
        # print(self.df_timeseries_stable)

        # PSVD accumulating over the stable time
        self.dict_psvd_stable = {key : val for key, val in self.dict_psvd_all.items()
                                    if self.start_stable_time <= key <= self.end_stable_time }
        #print(self.df_timeseries_all.shape, self.df_timeseries_stable.shape)
        if len(self.dict_psvd_stable) == 0:
            raise ValueError(f'>>> No stable data.')
        print(len(self.dict_psvd_all), len(self.dict_psvd_stable))

        # set rate of rainfall (precipitation intensity) for marshall palmer distribution
        if dict_precip is None:
            self.ser_precip_stable = None
            self.rate_rainfall = rate_rainfall
        else:
            self.ser_precip_stable = pd.Series({key : val for key, val in dict_precip.items()
                                                    if self.start_stable_time <= key <= self.end_stable_time })
            print(self.ser_precip_stable, self.ser_precip_stable.mean())
            self.rate_rainfall = self.ser_precip_stable.mean()

        # average PSVD per measurement
        self.df_psvd_stable = self._calc_df_psvd(self.dict_psvd_stable,
                                                    self.raindrop_diameters, self.raindrop_velocities)
        print(self.df_psvd_stable)

        # show the PSVD as heatmap
        if show_heatmap_p is not False:
            self._show_psvd_as_heatmap(self.df_psvd_stable)

        # stable timeseries of the rainfall measurements
        self.df_timeseries_stable = self._make_df_timeseries(self.dict_psvd_stable,
                                                                self.raindrop_diameters)
        print(self.df_timeseries_stable)

        # calc distribution statistics (mean and std)
        self._calc_stat()
        self.df_stat = self._make_df_stat()

        # make DSD (rainDrop Size Distribution, N(D)) as Pandas Series
        self.sr_dsd = self._make_dsd()
        print(self.sr_dsd)
        self.marshall_palmer = self.calc_marshall_palmer(self.rate_rainfall, self.raindrop_diameters)
        self.df_distribution_stable['DSD'] = self.sr_dsd
        self.df_distribution_stable['M-P'] = self.marshall_palmer

        # make PD (Particle Density) as Pandas Series
        self.sr_pd = self._make_pd()
        print(self.sr_pd)

    def _make_df_heatmap(self, df_psvd):
        _il = [i for i in df_psvd.index if i < 8]
        _cl = [i for i in df_psvd.columns if i < 11]
        _df_t = df_psvd.loc[_il,_cl]
        #print(_df_t)
        #print([i for i in _df_t.index if i < 8])
        #print([i for i in _df_t.columns if i < 10])

        return _df_t

    ##
    # @brief show the PSVD matrix per 10 sec as a heatmap
    def _show_psvd_as_heatmap(self, df_psvd):
        plt.clf() # reset pyplot graphs
        #sns.set(font_scale=0.5)
        #sns.heatmap(df_psvd.sort_index(axis=1,ascending=False).T.div(self.sampling_interval/10.0), vmin=0)
        plt.rcParams['figure.subplot.bottom'] = 0.175  # adjust pyplot bottom margins
        plt.rcParams['figure.subplot.top'] = 0.975
        
        _df_heatmap = self._make_df_heatmap(df_psvd)
        sns.heatmap(_df_heatmap.sort_index(axis=1,ascending=False).T.div(self.sampling_interval/10.0), vmin=0)

        #plt.rcParams['figure.subplot.left'] = 0.175
        #plt.rcParams['figure.subplot.right'] = 0.95
        plt.show()

    ##
    # @brief save the PSVD matrix per 10 sec as a heatmap image
    def save_psvd_as_heatmap(self, fbase, max_value=200):
        plt.clf()
        plt.rcParams['figure.subplot.top'] = 0.975
        plt.rcParams['figure.subplot.bottom'] = 0.175
        #sns.set(font_scale=0.8)
        _df_heatmap = self._make_df_heatmap(self.df_psvd_stable)
        #sns.heatmap(self.df_psvd_stable.sort_index(axis=1,ascending=False).T.div(self.sampling_interval/10.0),vmin=0,vmax=max_value)
        sns.heatmap(_df_heatmap.sort_index(axis=1,ascending=False).T.div(self.sampling_interval/10.0),vmin=0,vmax=max_value)
        #pngname = 'RSLT/' + fbase + '.png'
        pngname = fbase + '.png'
        print('>> Save: ', pngname)

        #plt.rcParams['figure.subplot.left'] = 0.175
        #plt.rcParams['figure.subplot.right'] = 0.95
        plt.savefig(pngname, dpi=300)

    def _extract_dims_velocs(self, dict_psvd):
        _df_psvd = list(dict_psvd.values())[0]
        return list(_df_psvd.index), list(_df_psvd.columns)

    def calc_diameter_differences(self, diams=None):
        if diams is None:
            diams = self.raindrop_diameters

        _l = [diams[0]]
        for i in range(1,len(diams)):
            _l.append(diams[i]-diams[i-1])

        return _l

    def calc_marshall_palmer(self, rate_rainfall, diams=None):
        if diams is None:
            diams = self.raindrop_diameters

        return [8000.0 * np.exp(-4.1*np.power(rate_rainfall,-0.21)*d) for d in diams]

    ##
    # @brief calculate a PSVD matrix on average per measurement
    #
    # @return the PSVD matrix as pandas DataFrame
    #
    def _calc_df_psvd(self, dict_psvd, diams=None, velocs=None):
        _psvd_stable = self._accum_df_psvd(dict_psvd, diams, velocs)

        return _psvd_stable / len(dict_psvd)

    ##
    # @brief make a PSVD matrix accumulating over stable rainfall time
    #
    # @return the PSVD matrix as pandas DataFrame
    #
    def _accum_df_psvd(self, dict_psvd, diams=None, velocs=None):
        if diams is None:
            diams = self.raindrop_diameters
        if velocs is None:
            velocs = self.raindrop_velocities

        df_psvd_stable = pd.DataFrame(index=diams, columns=velocs, data=0, dtype=int)
        df_psvd_stable.index.name = 'Diameter'
        df_psvd_stable.columns.name = 'Velocity'

        for _psvd in self.dict_psvd_stable.values():
            #print({_key:_psvd})
            df_psvd_stable += _psvd

        return df_psvd_stable

    ##
    # @brief make num of raindrops per size timeseries over stable rainfall time
    #
    # @return the num of raindrop per size timeseries as pandas DataFrame
    #
    """ Example of the num of raindrops per size timeseries
Time      2022-12-09 11:33:00  2022-12-09 11:33:10  2022-12-09 11:33:20  2022-12-09 11:33:30  ...  2022-12-09 11:40:30  2022-12-09 11:40:40  2022-12-09 11:40:50  2022-12-09 11:41:00
Diameter                                                                                      ...                                                                                    
0.062                       0                    0                    0                    0  ...                    0                    0                    0                    0
0.187                       0                    0                    0                    0  ...                    0                    0                    0                    0
0.312                      42                   44                   40                   55  ...                    9                   13                   17                    5
0.437                     163                  166                   76                  176  ...                   56                   26                   48                   48
0.562                     303                  326                  145                  333  ...                  199                  152                  201                  141
0.687                     405                  458                  331                  387  ...                  327                  328                  324                  243
0.812                     334                  326                  297                  342  ...                  346                  341                  307                  340
0.937                     233                  257                  234                  239  ...                  218                  200                  208                  198
1.062                     115                  157                  134                  127  ...                  104                  105                  116                  125
1.187                      62                   67                   70                   74  ...                   74                   64                   67                   76
1.375                      80                   83                   90                   84  ...                   70                   67                   63                   64
1.625                      36                   32                   39                   39  ...                   34                   16                   17                   32
1.875                       8                   10                    9                   21  ...                   10                   16                   13                   15
2.125                       6                    5                    9                    4  ...                    4                    7                    6                    2
2.375                       4                    5                    5                    4  ...                    4                    0                    7                    3
2.750                       6                    2                    6                    2  ...                    2                    2                    2                    2
3.250                       0                    1                    1                    2  ...                    0                    0                    0                    1
3.750                       0                    0                    0                    0  ...                    0                    0                    0                    0
4.250                       0                    0                    0                    0  ...                    1                    0                    0                    0
4.750                       0                    0                    0                    0  ...                    0                    0                    0                    0
5.500                       0                    0                    0                    0  ...                    0                    0                    0                    0
6.500                       0                    0                    0                    0  ...                    0                    0                    0                    0
7.500                       0                    0                    0                    0  ...                    0                    0                    0                    0
8.500                       0                    0                    0                    0  ...                    0                    0                    0                    0
9.500                       0                    0                    0                    0  ...                    0                    0                    0                    0
11.000                      0                    0                    0                    0  ...                    0                    0                    0                    0
13.000                      0                    0                    0                    0  ...                    0                    0                    0                    0
15.000                      0                    0                    0                    0  ...                    0                    0                    0                    0
17.000                      0                    0                    0                    0  ...                    0                    0                    0                    0
19.000                      0                    0                    0                    0  ...                    0                    0                    0                    0
21.500                      0                    0                    0                    0  ...                    0                    0                    0                    0
24.500                      0                    0                    0                    0  ...                    0                    0                    0                    0
    """
    def _make_df_timeseries(self, dict_psvd, diams=None):
        if diams is None:
            diams = self.raindrop_diameters

        df_timeseries_stable = pd.DataFrame(index=diams, dtype=int)
        df_timeseries_stable.index.name = 'Diameter'
        df_timeseries_stable.columns.name = 'Time'

        #for _time, _psvd in self.dict_psvd_stable.items():
        #    df_timeseries_stable[_time] = _psvd.sum(axis='columns')

        _l_ser = [df_timeseries_stable]
        for _time, _psvd in self.dict_psvd_stable.items():
            _ser = _psvd.sum(axis='columns')
            _ser.name = _time
            _l_ser.append(_ser)

        return pd.concat(_l_ser, axis=1)

    ##
    # @brief Make DSD (rainDrop Size Distribution, N_D \Delta D)
    #
    # N(D_i) = (1/At) \sum_{j} N_{ij}/V_j,
    # N_D \Delta D = N(D_i) / \Delta D_i,
    #  where N(D_i): num of raindrops par volume in size category i,
    #  A: sampling area [m^2], t: sampling time [sec],
    #  N_{ij}: total num of raindrops in size category i and speed category j,
    #  V_j: raindrop velocity in category j [m/sec], and
    #  \Delta D_i: width of size category i [mm]
    #
    # @return DSD as a pandas Series [m^{-3}mm^{-1}]
    #
    """ Example of the DSD
Diameter
0.062              NaN
0.187              NaN
0.312      1781.997946
0.437      5663.821151
0.562     12563.756108
0.687     19872.591992
0.812     17608.383283
0.937     11367.863930
1.062      6054.472269
1.187      3090.020675
1.375      1877.559668
1.625       540.627418
1.875       219.479147
2.125        97.199490
2.375        46.802025
2.750        23.727099
3.250         4.475654
3.750         1.162258
4.250         0.348287
4.750         0.237132
5.500              NaN
6.500              NaN
7.500              NaN
8.500              NaN
9.500              NaN
11.000             NaN
13.000             NaN
15.000             NaN
17.000             NaN
19.000             NaN
21.500             NaN
24.500             NaN
    """
    def _make_dsd(self):
        #_n_samplings = len(self.df_timeseries_stable.columns)
        #_sampling_volume = self.sampling_area * self.sampling_interval * _n_samplings
        _sampling_volume = self.sampling_area * self.sampling_interval

        _sr_NDi = (self.df_psvd_stable / self.raindrop_velocities).sum(axis='columns') / _sampling_volume
        #print(_sr_NDi)
        #print(self.raindrop_diameter_binwidth)
        _sr_ND = _sr_NDi / self.raindrop_diameter_binwidth
        #print(_sr_ND)

        return _sr_ND.where(_sr_ND>0.0, np.nan)

    def _make_pd(self):
        #_n_samplings = len(self.df_timeseries_stable.columns)
        #_sampling_volume = self.sampling_area * self.sampling_interval * _n_samplings
        _sampling_volume = self.sampling_area * self.sampling_interval

        _sr_NDi = (self.df_psvd_stable / self.raindrop_velocities).sum(axis='columns') / _sampling_volume
        #print(_sr_NDi)
        #print(self.raindrop_diameter_binwidth)
        #_sr_ND = _sr_NDi / self.raindrop_diameter_binwidth
        #print(_sr_ND)

        return _sr_NDi.where(_sr_NDi>0.0, np.nan)
    ##
    # @brief Set stable rainfall times
    # @note The start/end times, including unstable rainfall, and the year,
    #       month, and date have already been loaded from Excel.
    def _set_stable_times(self, start_stable_timestr,
                                end_stable_timestr,
                                start_stable_time_offset_min,
                                end_stable_time_offset_min):
        # the start/end times loaded from Excel
        #self.start_time = self.df_timeseries_all.columns[0]
        #self.end_time = self.df_timeseries_all.columns[-1]
        _datetimes = self.dict_psvd_all.keys()
        #print(_datetimes)
        self.start_time = min(_datetimes)
        self.end_time = max(_datetimes)
        print('>> All:', self.start_time, ' - ',  self.end_time)
        # default start/end times of stable rainfall
        self.start_stable_time = self.start_time
        self.end_stable_time = self.end_time
        _datestr = self.start_time.strftime('%Y-%m-%d')

        # they are overwritten, if specified
        if start_stable_timestr:
            _stime = pd.to_datetime(_datestr + ' ' + start_stable_timestr)
            #if (self.start_time > _stime):
            #    raise ValueError('>>> Start Time out of range')
            #else:
            if (self.start_time < _stime):
                self.start_stable_time = _stime

        if end_stable_timestr:
            _etime = pd.to_datetime(_datestr + ' ' + end_stable_timestr)
            #if (self.end_time < _etime):
            #    raise ValueError(f'>>> End Time out of range: {_etime}')
            #else:
            if (self.end_time > _etime):
                self.end_stable_time = _etime

        # and they are offset, if specified
        if start_stable_time_offset_min != 0:
            self.start_stable_time += offsets.Minute(start_stable_time_offset_min)
        if end_stable_time_offset_min != 0:
            self.end_stable_time += offsets.Minute(end_stable_time_offset_min)

        print('>> Stable:', self.start_stable_time, ' - ',  self.end_stable_time)

    def append(self, pda):
        #_df_data = pd.concat([self.df_timeseries_all, pda.df_timeseries_all], axis=1)
        #self.df_timeseries_all = _df_data

        _df_stable = pd.concat([self.df_timeseries_stable, pda.df_timeseries_stable], axis=1)
        self.df_timeseries_stable = _df_stable

        self.dict_psvd_all.update(pda.dict_psvd_all)
        self.dict_psvd_stable.update(pda.dict_psvd_stable)
        #self.df_psvd_stable += pda.df_psvd_stable
        self.df_psvd_stable = self._calc_df_psvd(self.dict_psvd_stable,
                                                    self.raindrop_diameters, self.raindrop_velocities)

        self._calc_stat()
        self.df_stat = self._make_df_stat()

        if self.ser_precip_stable is None:
            self.ser_precip_stable = pda.ser_precip_stable
        else:
            if pda.ser_precip_stable is not None:
                _ser_precip = pd.concat([self.ser_precip_stable, pda.ser_precip_stable])
                self.ser_precip_stable = _ser_precip

        if self.ser_precip_stable is not None:
            self.rate_rainfall = self.ser_precip_stable.mean()

        self.sr_dsd = self._make_dsd()
        self.df_distribution_stable['DSD'] = self.sr_dsd
        #self.df_distribution_stable['MP'+str(self.rate_rainfall)] = self.MARSHALL_PALMER
        self.df_distribution_stable['M-P'] = self.marshall_palmer

    def _calc_stat(self):
        _dft_stable = self.df_timeseries_stable.T

        # create dataframe for mean number of particles
        _ser_sum = _dft_stable.sum(axis=0)
        _ser_mean = _dft_stable.mean(axis=0)
        self.df_distribution_stable = pd.DataFrame(index=self.raindrop_diameters, dtype=int)
        self.df_distribution_stable.index.name = 'Diameter'
        #ser_std = dft_range.std(axis=0, ddof=0) # p-std (1/N)
        self.df_distribution_stable['Sum'] = _ser_sum
        self.df_distribution_stable['Mean'] = _ser_mean
        self.df_distribution_stable['Mean/10sec'] = _ser_mean / (self.sampling_interval/10.0)
        #self.df_timeseries_stable['Sum'] = ser_sum
        #self.df_timeseries_stable['Mean'] = ser_mean
        #df_stable['Std.P'] = ser_std

        # calc mean and std of particle diameter
        #dft_stable = self.df_timeseries_stable.T

        _idx_diam = _dft_stable.columns
        self.total_num = _ser_sum.sum()
        self.total_mean = (_ser_sum * _idx_diam).sum() / self.total_num
        _total_diam2 = (_ser_sum * _idx_diam * _idx_diam).sum()
        _total_sigma = _total_diam2 / self.total_num - self.total_mean * self.total_mean
        self.total_std = np.sqrt(_total_sigma)   # p-std (1/N)

        #self.max_mean_value = _ser_mean.max()
        self.max_mean_value = self.df_distribution_stable['Mean/10sec'].max()

        #print('Mean: ', self.total_mean)
        #print('Std: ', self.total_std)

    def _make_df_stat(self):
        #_imval = int(self.max_mean_value) + 20
        _imval = 1680 # fixed value for report (2024/01/11)
        _imval -= _imval%10  # Y-scale is max value + 20, rounded down to the nearest ten
        _data = {0:[self.total_mean,self.total_mean-self.total_std,self.total_mean+self.total_std],
                    _imval:[self.total_mean,self.total_mean-self.total_std,self.total_mean+self.total_std]}
        df_stat = pd.DataFrame(data=_data, index=['Mean', '-Sigma', '+Sigma'])

        return df_stat

    def save(self, xlsxname, sname_base='Sheet',
                sname_stable='Sheet-Stable', sname_distribution='Sheet-Distribution',
                sname_psvd='Sheet-PSVD', sname_stat='Sheet-Stat'):
        #_dft_all = self.df_timeseries_all.T
        _dft_stable = self.df_timeseries_stable.T
        _dft_stable.index = _dft_stable.index.tz_localize(None) # Because DateTime with timezone cannot be saved in Excel.
        _dft_distribution = self.df_distribution_stable.T
        _dft_psvd = self.df_psvd_stable.T
        with pd.ExcelWriter(xlsxname) as writer:
            #_dft_all.to_excel(writer, sheet_name=sname_base)
            _dft_stable.to_excel(writer, sheet_name=sname_stable)
            _dft_distribution.to_excel(writer, sheet_name=sname_distribution)
            _dft_psvd.to_excel(writer, sheet_name=sname_psvd)
            self.df_stat.to_excel(writer, sheet_name=sname_stat)

    def _create_scatterchart_raw(self, ws, stat_sheet, max_y_value=None, chart_title=None):
        chart = ScatterChart()

        # chart elements
        chart.title = chart_title
        chart.x_axis.title = "Particle Diameter [mm]"
        chart.y_axis.title = "Num of Particles [N/10sec]"
        chart.legend = None
        chart.width = 20
        chart.height = 15

        # graph of the raindrop diameters
        #   row[1]: Size, row[2]:Sum, row[3]: Mean, row[4]:Mean/10sec, row[5]:DSD, row[6]:M-P
        x_values = Reference(ws, min_row=1, max_row=1, min_col=2, max_col=ws.max_column)
        y_values = Reference(ws, min_row=4, max_row=4, min_col=2, max_col=ws.max_column)

        series = Series(y_values, x_values, title=chart_title, title_from_data=False)
        series.graphicalProperties.line.solidFill = "4A7EBB"
        series.marker.symbol ="square"
        series.marker.size = 5
        series.marker.graphicalProperties.solidFill = "4A7EBB"
        series.marker.graphicalProperties.line.solidFill = "4A7EBB"
        chart.series.append(series)

        # vertical line showing the mean value
        x_mean = Reference(stat_sheet, min_row=2, max_row=2, min_col=2, max_col=3)
        y_height = Reference(stat_sheet, min_row=1, max_row=1, min_col=2, max_col=3)

        series_mean = Series(y_height, x_mean, title='mean', title_from_data=False)
        series_mean.graphicalProperties.line.solidFill = "00B0F0"
        series_mean.graphicalProperties.line.dashStyle = "dash"
        series_mean.graphicalProperties.line.width = pixels_to_EMU(1.5)
        lb = DataLabel(idx=1, showSerName=True, showVal=False)
        lbl = DataLabelList(dLbl=[lb])
        series_mean.labels = lbl
        chart.series.append(series_mean)

        # vertical line showing the mean - std value
        x_msigma = Reference(stat_sheet, min_row=3, max_row=3, min_col=2, max_col=3)
        series_msigma = Series(y_height, x_msigma, title='-sigma', title_from_data=False)
        series_msigma.graphicalProperties.line.solidFill = "00B0F0"
        series_msigma.graphicalProperties.line.dashStyle = "dash"
        series_msigma.graphicalProperties.line.width = pixels_to_EMU(1.5)
        series_msigma.labels = lbl
        chart.series.append(series_msigma)

        # vertical line showing the mean + std value
        x_psigma = Reference(stat_sheet, min_row=4, max_row=4, min_col=2, max_col=3)
        series_psigma = Series(y_height, x_psigma, title='+sigma', title_from_data=False)
        series_psigma.graphicalProperties.line.solidFill = "00B0F0"
        series_psigma.graphicalProperties.line.dashStyle = "dash"
        series_psigma.graphicalProperties.line.width = pixels_to_EMU(1.5)
        series_psigma.labels = lbl
        chart.series.append(series_psigma)

        # chart size
        chart.x_axis.scaling.min = 0
        chart.x_axis.scaling.max = 6
        chart.y_axis.scaling.min = 0
        #chart.y_axis.scaling.max = 200
        #imval = int(self.max_mean_value) + 20
        imval = int(max_y_value) + 20
        chart.y_axis.scaling.max = imval - imval%10
        #chart.y_axis.scaling.max = 1800

        return chart

    def _create_scatterchart_dsd(self, ws, rate_rainfall, chart_title=None):
        chart = ScatterChart()

        # chart elements
        chart.title = chart_title
        chart.x_axis.title = "Particle Diameter [mm]"
        chart.y_axis.title = "Num of Particles [m-3 mm-1]"
        #chart.x_axis.axPos = 't'
        #chart.x_axis.lblAlgn = 'l'
        chart.x_axis.tickLblPos = 'low'
        chart.legend = None
        chart.width = 20
        chart.height = 15

        # chart size
        chart.x_axis.scaling.min = 0
        chart.x_axis.scaling.max = 10
        chart.y_axis.scaling.min = 0.01
        #chart.y_axis.scaling.max = 200
        #imval = int(self.max_mean_value) + 20
        #chart.y_axis.scaling.max = imval - imval%10
        chart.y_axis.scaling.logBase = 10
        #chart.y_axis.majorGridlines = None

        # graph of the raindrop diameters
        x_values = Reference(ws, min_row=1, max_row=1, min_col=2, max_col=ws.max_column)
        y_values = Reference(ws, min_row=5, max_row=5, min_col=2, max_col=ws.max_column)

        series = Series(y_values, xvalues=x_values, title=chart_title, title_from_data=False)
        series.graphicalProperties.line.solidFill = "4A7EBB"
        series.marker.symbol ="square"
        series.marker.size = 5
        series.marker.graphicalProperties.solidFill = "4A7EBB"
        series.marker.graphicalProperties.line.solidFill = "4A7EBB"
        chart.series.append(series)

        # draw Marshall-Palmer distribution
        mp_values = Reference(ws, min_row=6, max_row=6, min_col=2, max_col=ws.max_column)
        series_mp = Series(mp_values, x_values, title='M-P '+'{:.3g}'.format(rate_rainfall), title_from_data=False)
        series_mp.graphicalProperties.line.solidFill = "00B0F0"
        series_mp.graphicalProperties.line.dashStyle = "dash"
        series_mp.graphicalProperties.line.width = pixels_to_EMU(1.5)
        lb = DataLabel(idx=1, showSerName=True, showVal=False)
        lbl = DataLabelList(dLbl=[lb])
        series_mp.labels = lbl
        chart.series.append(series_mp)

        return chart

    def save_xlsx_with_graph(self, fbase):
        try:
            with tempfile.TemporaryDirectory() as dname:
                #_tmpfile = dname + '/' + fbase + '.xlsx'
                _tmpfile = dname + 'tmp.xlsx'
                #print(_tmpfile)
                self.save(_tmpfile, fbase, fbase+'-Stable', fbase+'-Distribution', fbase+'-PSVD', fbase+'-Stat')
                #print('Reload: ', _tmpfile)
                wb = load_workbook(_tmpfile)

            #ws_stable = wb[fbase+'-Stable']
            ws_dist = wb[fbase+'-Distribution']
            ws_stat = wb[fbase+'-Stat']

            # lchart = pda.create_linechart(ws)
            #print('Create ScatterChart')
            #ws_stat = wb.create_sheet(fbase+'-Stat')
            #dsd.fill_stat_sheet(ws_stat)
            #cchart = self._create_scatterchart_raw(ws_dist, ws_stat, self.max_mean_value, fbase)
            cchart = self._create_scatterchart_raw(ws_dist, ws_stat, max_y_value=1700, chart_title=fbase)  # MAX y value fixed for report (2024/01/11)
            dchart = self._create_scatterchart_dsd(ws_dist, self.rate_rainfall, fbase)

            ws_graph = wb.create_sheet(fbase+'-Graph')

            #graph_sheet.add_chart(lchart, f"B2")
            ws_graph.add_chart(cchart, f"B2")
            ws_graph.add_chart(dchart, f"N2")

            xlsxname = fbase + '.xlsx'
            print('>> Save: ', xlsxname)
            wb.save(xlsxname)

        except Exception as e:
            print(e)
